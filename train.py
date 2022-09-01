"""
Created on Sept 11, 2020

train  IMRec model

@author: Shiying Ni
"""

import os
from datetime import datetime
import time
import argparse
import json
from collections import defaultdict

# ============================= Config ============================

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

parser = argparse.ArgumentParser(description='Train model')
# Exp Params
parser.add_argument('--use_gpu', default=0, type=int)
parser.add_argument('--dataset', default='taobao')
parser.add_argument('--mode', default='train_500K')
parser.add_argument('--model_type', default='IMRec')
parser.add_argument('--run_times', default=1, type=int)  # repeated times
parser.add_argument('--K-int-type', nargs='+', default=[1, 5, 10, 20, 50], type=int)
parser.add_argument('--total_epochs', default=30, type=int)
parser.add_argument('--eval_interval', default=1, type=int)
parser.add_argument('--verbose', default=1, type=int)
parser.add_argument('--patience', default=3, type=int)
parser.add_argument('--save_weights', default=1, type=int)  # whether to save ckpts when training
parser.add_argument('--weights_dir', default='', type=str)  # load weights from ckpts
parser.add_argument('--learning_rate', default=-1, type=float)

# Model Params
parser.add_argument('--batch_size', default=-1, type=int)
parser.add_argument('--embed_dim', default=-1, type=int)
parser.add_argument('--att_len', default=-1, type=int)
parser.add_argument('--alpha', default=-1, type=float)
parser.add_argument('--without_il', default=-1, type=int)
parser.add_argument('--item_intention', default=-1, type=int)
parser.add_argument('--time_threshold', default=-1, type=int)  # 0 means do not set threshold
parser.add_argument('--bpr', default=-1, type=int)

args, unknown = parser.parse_known_args()

if not args.use_gpu:
    os.environ['CUDA_VISIBLE_DEVICES'] = '-1'
    print('Ban GPU device.')

import tensorflow as tf
import pandas as pd
import numpy as np
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import EarlyStopping, ModelCheckpoint
import tensorflow.keras.backend as backend
from openpyxl import load_workbook

from module import RecordLoss, RecordMetrics, HR, NDCG, MRR
from utils import get_data
from configs import _model_config, _ds_config, _exp_config, _models

# =============================== GPU ==============================
gpus = tf.config.experimental.list_physical_devices(device_type='GPU')
print("Num GPUs Available:", len(gpus))
if gpus:
    tf.config.experimental.set_visible_devices(gpus[:], 'GPU')
    for gpu in gpus:
        tf.config.experimental.set_memory_growth(gpu, True)
    print('Use GPU when training.')
else:
    print('Do not use GPU when training.')

# =========================Hyper Parameters =======================
# key params
dataset = args.dataset
mode = args.mode
model_type = args.model_type

ds_config = _ds_config.get(model=model_type, dataset=dataset, mode=mode)
model_config = _model_config.get(model=model_type, dataset=dataset, mode=mode)
exp_config = _exp_config.get(model=model_type, dataset=dataset, mode=mode)

# experimental params
run_times = args.run_times
total_epochs = args.total_epochs
eval_interval = args.eval_interval
verbose = args.verbose
K = args.K_int_type
patience = args.patience
learning_rate = args.learning_rate if args.learning_rate > -1 else exp_config['learning_rate']
batch_size = args.batch_size if args.batch_size > -1 else exp_config['batch_size']

# dataset params
embed_dim = args.embed_dim if args.embed_dim > -1 else ds_config['embed_dim']
item_intention = args.item_intention if args.item_intention > -1 else ds_config['item_intention']
time_threshold = args.time_threshold if args.time_threshold > -1 else model_config['time_threshold']
with_ts = True if time_threshold else False
# update dataset params
ds_config.update(
    embed_dim=embed_dim,
    item_intention=item_intention,
    with_ts=with_ts,
    )

# IMRec params
if model_type == 'IMRec':
    att_len = args.att_len if args.att_len > -1 else model_config['att_len']
    alpha = args.alpha if args.alpha > -1 else model_config['alpha']
    bpr = args.bpr if args.bpr > -1 else model_config['BPR']
    without_il = args.without_il if args.without_il > -1 else model_config['without_il']
    model_config.update(
        att_len=att_len, alpha=alpha, without_il=without_il, BPR=bpr,
        time_threshold=time_threshold,
        )

exp_params = {}
print('Experimental Params'.center(50, '='))
for arg in vars(args):
    if getattr(args, arg) != -1:
        print(f'{arg.capitalize()}: {getattr(args, arg)}')
        exp_params[arg] = getattr(args, arg)


# ========================== Create dataset =======================
data_info, train, val, test = get_data(
    dataset=dataset,
    mode=mode,
    regenerate=False,
    **ds_config,
    )

# dataset params
print('Dataset Params'.center(50, '='))
for key, value in ds_config.items():
    print(f'{key.capitalize()}: {value}')

def generate_dataset(data):
    label = data['target_item_pos']  # dummy ouputs
    return data, label

trainsets, trainlabel = generate_dataset(train)
validationsets, validationlabel = generate_dataset(val)
testsets, testlabel = generate_dataset(test)

now = datetime.now().strftime('%m%d%H%M')
#%%
# ===========================Run Experiments============================
if __name__ == '__main__':

    summary_list = []
    exp_begin = time.time()
    for run in range(1, run_times + 1):
        model = None

    # =============================Build Model==============================
        model = _models[model_type](data_info, model_config)

        val_metrics = 'MRR'
        val_metrics_func = [MRR()]

        model.compile(
            optimizer=Adam(learning_rate=learning_rate),
            metrics=val_metrics_func,
            )

        log_dir = './log'
        os.makedirs(log_dir, exist_ok=True)

        model_name = model.model_name
        para_info = f'{now}_{model_name}_{dataset}-{mode[6:]}_bs_{batch_size}_dim_{embed_dim}_neg_{ds_config["test_neg_ratio"]}'

        # save model
        if args.model_type == 'IMRec':
            para_info += f'_len_{att_len}'
            para_info += f'_a_{alpha}'
            if item_intention:
                para_info += '_II'

        exp_dir = f'{log_dir}/{para_info}'
        os.makedirs(exp_dir, exist_ok=True)

        # model params
        print('Model Params'.center(50, '='))
        for key, value in model_config.items():
            print(f'{key.capitalize()}: {value}')

        # ========================Fit and Evaluate============================
        print(f'{model.model_name}: {run}/{run_times} run'.center(50, '*'))

        results_list = []
        best_score = 0.0
        best_model = None
        loss_log = defaultdict(list)
        metrics_log = defaultdict(list)
        ckpt_dir = f'{exp_dir}/run_{run}'
        early_stopping = EarlyStopping(
            monitor=f'val_{val_metrics}',
            min_delta=0.0001,
            patience=patience,
            verbose=1,
            mode='max',
            baseline=None,
            restore_best_weights=True)
        callbacks = [
            RecordLoss(loss_log, mode='batch'),  # record loss in each batch
            RecordMetrics(testsets, metrics_log, interval=eval_interval, K=K),  # record metrics on the testsets
            early_stopping,
                     ]
        if args.save_weights:
            model_checkpoint = ModelCheckpoint(
                filepath=ckpt_dir+'/best_epoch.ckpt',
                save_weights_only=True,
                monitor=f'val_{val_metrics}',
                min_delta=0.0001,
                verbose=0,
                save_best_only=True,
                mode='max',
                )
            callbacks.append(model_checkpoint)

        if os.path.exists(os.path.dirname(args.weights_dir)):
            model.load_weights(args.weights_dir)
            print(f'Load weights from {args.weights_dir}.')

        model.fit(
            x=trainsets,
            y=trainlabel,
            validation_data=(validationsets, validationlabel),
            batch_size=batch_size,
            epochs=total_epochs,
            verbose=verbose,
            shuffle=True,
            callbacks=callbacks,
                  )

        metrics_this_run = pd.DataFrame(metrics_log)
        loss_this_run = pd.DataFrame(loss_log)
        loss_cols = [f'val_{val_metrics}'] + [col for col in loss_this_run.columns if 'loss' in col]
        metrics_loss_this_run = loss_this_run[['Epoch', 'Fitting Time'] + loss_cols]
        metrics_loss_this_run = metrics_loss_this_run.drop_duplicates(subset=['Epoch'], keep='last')  # 选择每个epoch最后一个batch的loss值
        metrics_this_run = pd.merge(metrics_this_run, metrics_loss_this_run, on='Epoch', how='left')

        if early_stopping.stopped_epoch == 0:
            best_epoch = total_epochs - 1
        else:
            best_epoch = early_stopping.stopped_epoch - patience

        best_metrics = metrics_this_run.loc[best_epoch:best_epoch, :].copy()
        best_metrics.loc[best_epoch, 'Run'] = run
        summary_list.append(best_metrics)

        # =========================Save Model and Write Log Files=======================
        # save loss
        loss_file = exp_dir + f'/loss_{para_info}.xlsx'
        # create new .xlsx file
        if not os.path.exists(loss_file):
            pd.DataFrame([]).to_excel(loss_file,
                                      sheet_name='summary',
                                      header=True, index=False)
        # add loss sheet for each run
        book = load_workbook(loss_file)
        with pd.ExcelWriter(loss_file, engine='openpyxl') as excel_writer:
            excel_writer.book = book
            excel_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            loss_this_run.to_excel(excel_writer,
                                   sheet_name='{}th_run'.format(run),
                                   header=True, index=False)
            print(f'Save loss of {run}th run'.center(50, '*'))

        # save metrics
        metrics_file = exp_dir + f'/metrics_{para_info}.xlsx'
        # create new .xlsx file
        if not os.path.exists(metrics_file):
            pd.DataFrame([]).to_excel(metrics_file,
                                      sheet_name='summary',
                                      header=True, index=False)
        # add metrics sheet for each run
        book = load_workbook(metrics_file)
        with pd.ExcelWriter(metrics_file, engine='openpyxl') as excel_writer:
            excel_writer.book = book
            excel_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            metrics_this_run.to_excel(excel_writer,
                                      sheet_name='{}th_run'.format(run),
                                      header=True, index=False)
            print(f'Save metrics of {run}th run'.center(50, '*'))

        # save summary of this exp
        summary = pd.concat(summary_list)
        summary.loc['average'] = summary.mean()
        summary.loc['average', 'Run'] = 'average'
        with pd.ExcelWriter(metrics_file, engine='openpyxl') as excel_writer:
            excel_writer.book = book
            excel_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            summary.to_excel(excel_writer,
                             sheet_name='summary',
                             header=True, index=True)

    # save params
    params = {}
    params.update(exp_params)
    params.update(exp_config)  # experiment params
    params.update(ds_config)  # dataset params
    params.update(model_config)  # model params

    trainable_count = int(np.sum([backend.count_params(w) for w in model.trainable_weights]))
    non_trainable_count = int(np.sum([backend.count_params(w) for w in model.non_trainable_weights]))
    params['trainable_params'] = trainable_count
    params['non_trainable_params'] = non_trainable_count

    with open(f'{exp_dir}/params.json', 'a') as file:
        json.dump(params, file)

    # after the exp, identity it with best_score
    best_score = summary.loc['average', val_metrics]
    new_exp_dir = f'{exp_dir}_{best_score:.4f}'
    os.rename(src=exp_dir, dst=new_exp_dir)

    print(f'Rename exp dir as {new_exp_dir}')
    exp_end = time.time()
    total_time = (exp_end - exp_begin) / 60
    time_per_run =  total_time / run_times

    print('Experiment done'.center(50, '='))

    print('Trainable params: {:,}'.format(int(trainable_count)))
    print('Non-trainable params: {:,}'.format(int(non_trainable_count)))
    print(f'Average best score of {run_times} runs: {val_metrics} = {best_score:.4f}')

    print(f'{run_times}-run time: {total_time:.2f} min.')
    print(f'Average time: {time_per_run:.2f} min.')
